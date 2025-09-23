import easyocr
import cv2
import numpy as np
import os
from langchain_openai import ChatOpenAI
from langchain.agents import initialize_agent, AgentType
from langchain.prompts import PromptTemplate
from langchain.tools import tool
from langchain.tools import StructuredTool
from pydantic import BaseModel, Field, field_validator
from typing import List
from openpyxl import Workbook
from datetime import date
from dotenv import load_dotenv
import sys

def mainFunction(byteImage: bytes):
    # add path of parent dir
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    if project_root not in sys.path:
        sys.path.append(project_root)
    
    try:
        load_dotenv()
    except:
        pass

    # debugging purpose
    # if not os.getenv("LANGSMITH_API_KEY"):
    #     print("LANGSMITH_API_KEY_NOT_FOUND")
    # if not os.getenv("LANGSMITH_PROJECT"):
    #     print("LANGSMITH_PROJECT is set to default")
    #     os.environ["LANGSMITH_PROJECT"] = "default"
    # if not os.getenv("LANGSMITH_TRACING_V2"):
    #     print("LANGSMITH_TRACING_V2 is set to true")
    #     os.environ["LANGSMITH_TRACING_V2"] = "true"
        
    if not os.getenv("OPENAI_API_KEY"):
        print("OPENAI_API_KEY_NOT_FOUND")
        exit(1)
    else:
        openai_valid = True

    # languages that ocr will read.
    reader = easyocr.Reader(['en'])
    current_dir = os.path.dirname(__file__)

    # image of receipt
    image_path = os.path.join(current_dir, "sample.jpg")
    np_arr = np.frombuffer(byteImage, np.uint8)
    image = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    # image = cv2.imread(image_path)
    results = reader.readtext(image)

    # 1. align based on y-axis
    results.sort(key=lambda x: (x[0][0][1], x[0][0][0]))  # x[0][0][1] = top-left y pos. then top-left x pos. upper one first, if the same, left one first.s

    # 2. combine same row words into a single array
    line_threshold = 10  # assume as the same row if y-diff is less than 10px
    lines = []
    current_line = []
    last_y = -1*line_threshold
    for bbox, text, conf in results:
        # pts = np.array(bbox, np.int32) # debugging purpose
        # cv2.polylines(image, [pts], True, (0,255,0), 2) # debugging purpose
        y = bbox[0][1]  # top-left y
        if not current_line:
            current_line.append((bbox, text, conf))
        elif abs(y - last_y) <= line_threshold:
            current_line.append((bbox, text, conf))
        else:
            lines.append(current_line)
            current_line = [(bbox, text, conf)]
        last_y = y

    # add last line
    if current_line:
        lines.append(current_line)
        
    # cv2.imshow("Amounts Highlighted", image) # debugging purpose
    # cv2.waitKey(0) # debugging purpose

    # 3. check
    temp_text = ""
    for i, line in enumerate(lines):
        temp_text += f"Line {i+1}. "
        temp_text += (" | ".join([t[1] for t in line]))
        temp_text += "\n"

    # print(f" {line_text}")
    ########################################
    ##        picture parsing done        ##
    ########################################

    class Receipt(BaseModel):
        name:str = Field(description="Item name")
        count:int = Field(description="Item count")
        price:float = Field(description="This item's total price. Not a price of a each item")

    class ItemList(BaseModel):
        items: List[Receipt] = Field(description="total item information. each item contain it's name, count, and own price.")
        total_price: float = Field(description="total price of entire items. Also should be found from user input.")
        # bought_date: date = Field(description="bought date.") #debugging
        
        # @field_validator('bought_date', mode='before')
        # def parse_date(cls, v):
        #     if isinstance(v, str):
        #         v = v.replace("/", "-")
        #     return v
        
    epsilon = 1e-5
    def floating_point_error(a:float, b:float):
        if abs(a-b) < epsilon:
            return False
        return True 

    # tool that the agentic ai will call
    def Items_on_List(receipt: ItemList):
        # currently inputting the data into an excel file, but can be changed to others!
        wb = Workbook()
        
        ws=wb.active
        ws.title="receipt"
        ws["A2"] = "item name"
        ws["B2"] = "Number of items"
        ws["C2"] = "price of items"
        # ws["A1"] = f"{receipt.bought_date}"
        price_check=0
        for i, r in enumerate(receipt.items):
            ws[f"A{i+3}"] = f"{r.name}"
            ws[f"B{i+3}"] = r.count
            ws[f"C{i+3}"] = r.price
            price_check+=r.price

        ws["E2"] = "Total Cost"
        ws["E3"] = receipt.total_price
        # floating_point_error to see if the difference is bigger than epsilon.
        # if error is found, shows how much cost could be omitted. does not work well if receipt.total_price itself is wrong.
        if floating_point_error(receipt.total_price, price_check):
            print(f"Some items can be omitted. Please check. Total price omitted: {receipt.total_price - price_check}")
            
        # save as the excel file.
        file_path = os.path.join(current_dir, "receipt.xlsx")
        wb.save(file_path)
        return



    ########################################
    ##        receipt class done          ##
    ########################################

    model = ChatOpenAI(model='gpt-4.1-nano', temperature=0)

    tool = StructuredTool.from_function(
            name="organize_receipt",
            func=Items_on_List,
            description="receipt organizing tool"
        )

    # tool list for agent
    tools = [
        tool
    ]

    agent = initialize_agent(
        tools=tools,
        llm=model,
        agent=AgentType.OPENAI_FUNCTIONS
    )
    # llm_with_tool = model.bind_tools(tools)

    prompt = PromptTemplate.from_template("Can you organize this receipt? {receipt}")
    Actual_promt = prompt.invoke({"receipt": temp_text})
    # print(Actual_promt)

    # chain = LLMChain(llm=model, prompt=prompt)
    # output = model.invoke(Actual_promt)
    # output=llm_with_tool.invoke(Actual_promt)
    output = agent.run(Actual_promt)
    # chain.run(receipt=temp_text)
    # print(output.content)
    # print(output)
    return output
    
